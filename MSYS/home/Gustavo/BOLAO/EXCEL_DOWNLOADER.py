from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import yaml
from openpyxl import load_workbook
import openpyxl

import gspread.utils

def get_values(spreadsheet_id, range_name):
	credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)

    # pylint: disable=maybe-no-member
	try:
		service = build('sheets', 'v4', credentials=credentials)

		result = service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range=range_name).execute()
		rows = result.get('values', [])
		print(f"{len(rows)} rows retrieved")
		return rows
	except HttpError as error:
		print(f"An error occurred: {error}")
		return error

def batch_update_values(spreadsheet_id, data, value_input_option):
	credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)

	try:

		service = build('sheets', 'v4', credentials=credentials)



		body = {
            'valueInputOption': value_input_option,
            'data': data
        }

		result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id, body=body).execute()


		print(f"{result.get('totalUpdatedCells')} cells updated.")
		return result
	except HttpError as error:
		print(f"An error occurred: {error}")
		return error


def is_number(string):
    return string.isdigit()


with open('Configuracoes.yml', 'r') as f:
	data = yaml.load(f, Loader=yaml.FullLoader)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SERVICE_ACCOUNT_FILE = 'C:/Users/nicol/Downloads/Bolao_2023_24/bolao-premier-league-5a52cb742830.json'

with open('info_spreadsheet.yaml', 'r') as f:
	data1 = yaml.load(f, Loader=yaml.FullLoader)

with open('name_equivalence.yml', 'r') as f:
	data0 = yaml.load(f, Loader=yaml.FullLoader)

workbook_name = data['ARQUIVO_EXCEL_ABRIR']
workbook_future_name = data['ARQUIVO_EXCEL_SALVAR']
spreadsheet_id = data1['id']
range_name = 'Form Responses 1!A1:BM100'

array = get_values (spreadsheet_id, range_name)
# print(array)

# print(array)


# data =[]
infos = []
table = []
branco = None
temp = "a"
workbook = load_workbook(filename = workbook_name, read_only=False,keep_vba=True)


for row in range(len(array)):
	infos = []
	if row == 0:
		for col in range(len(array[row])):
			table_temp = []
			if col <= 2:
				continue
			elif (array[row][col] == "Email Address" or array[row][col] == "Endereço de e-mail"):
				continue
			elif col % 2 == 1:
				temp = array[row][col].split(" - P")
				# print(temp)
				peso = int(temp[1][0])
				# print(peso)
				temp = array[row][col].split(" [")
				time1 = temp[1].replace("]", "")
				temp = array[row][col+1].split(" [")
				time2 = temp[1].replace("]", "")
				lista = [time1, branco, "x", branco, time2, peso]
				table.append(lista)
		sheet = workbook['Tabela']
		for x in range(5000):
			if sheet[('A'+str(x+1))].value == data['Nome_Rodada']:
				position = ('A'+str(x+1))
		start_row = (int(position[1:])+1)
		start_column = openpyxl.utils.column_index_from_string("A")
		# print(start_row)
		# print(start_column)
		for i, value in enumerate(table):
			for i1 in range(len(table[i])):

				cell = sheet.cell(row=(start_row+i), column=(start_column+i1))
				cell.value = table[i][i1]

		continue
	for col in range(len(array[row])):
		if col < (len(array[row])-1):
			infos.append(array[row][col])

	# print(f"{data0[infos[0]]}")
	# print(f"{data0[infos[0]].rstrip()}")
	sheet = workbook[f"{data0[(infos[1].rstrip())]}"]
	for x in range(5000):
		if sheet[('A'+str(x+1))].value == data['Nome_Rodada']:
			position = ('A'+str(x+1))
	start_row = int(position[1:])
	start_column = openpyxl.utils.column_index_from_string("B")

	for i, value in enumerate(infos):
		cell = sheet.cell(row=start_row, column=start_column+i)
		if is_number(value) == True:
			cell.value = int(value)
		else:
			cell.value = (value)

	# print(sheet[position].value)
	# print(sheet[position])

workbook.save(workbook_future_name)