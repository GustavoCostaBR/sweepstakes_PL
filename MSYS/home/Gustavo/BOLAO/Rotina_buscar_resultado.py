# import var_dump
import subprocess
import yaml
from datetime import datetime, timedelta
import pytz
import sqlite3
import openpyxl
from openpyxl import load_workbook
# import pdb
# import sys

with open('Configuracoes.yml', 'r') as f:
	data = yaml.load(f, Loader=yaml.FullLoader)

with open('Ativar_Desativar_campeonatos.yml', 'r') as f:
	data0 = yaml.load(f, Loader=yaml.FullLoader)

with open('tabelas.yml', 'r') as f:
	data1 = yaml.load(f, Loader=yaml.FullLoader)

gmt = pytz.timezone('GMT')
dt = []
dt2 = []

# sys.stdout.reconfigure(encoding='utf-8')

def after_FACS_filter(variavel, inicio_bolao_date):

	current_year = datetime.now().year
	horarios_filtrados = []
	horarios_filtrados_date_time = []
	index_delete=[]
	# pdb.set_trace()
	for index in range(len(variavel)):
		horarios_filtrados.append(variavel[index].split(" - ")[2].strip())
		horarios_filtrados_date_time.append(datetime.strptime(horarios_filtrados[index], "%d/%m"))
		horarios_filtrados_date_time[index] = horarios_filtrados_date_time[index].replace(year=current_year)
		horarios_filtrados_date_time[index] = gmt.localize(horarios_filtrados_date_time[index])
		if horarios_filtrados_date_time[index] < inicio_bolao_date:
			print(horarios_filtrados_date_time[index])
			print(inicio_bolao_date)
			index_delete.append(index)

	variavel = [value for index, value in enumerate(variavel) if index not in index_delete]
	return variavel

def subst_P(peso, peso_replace, temp):
	index1 = 0
	if len(temp) > 0:
		while peso in temp[index1]:
			temp[index1] = temp[index1].replace(peso, peso_replace)
			if index1 < (len(temp)-1):
				index1=index1 + 1
	return temp

for numeros in data['Inicio_Rodada']:
	dt.append(str(numeros))

for numeros in data['Fim_Rodada']:
	dt2.append(str(numeros))

# dt = datetime(*data['Inicio_Rodada'], tzinfo=gmt)
# dt2 = datetime(*data['Fim_Rodada'], tzinfo=gmt)
url_FACS = data['URL_FACS']
url_PL = data['URL_PL']
url_SPFL = data['URL_SPFL']
url_SPFL_championship = data['URL_SPFL_CHAMPIONSHIP']
url_SPFL_relegation = data['URL_SPFL_RELEGATION']
url_SPFL_PLAYOFFS = data['URL_SPFL_PLAYOFFS']
url_LCH = data['URL_LCH']
url_LCH_PLAYOFFS = data['URL_LCH_PLAYOFFS']
url_FAC = data['URL_FAC']
url_EFL = data['URL_EFL']
url_SC = data['URL_SC']
url_SLC = data['URL_SLC']
url_USC = data['URL_USC']
url_QUAL_UCL = data['URL_QUAL_UCL']
url_UCL = data['URL_UCL']
url_QUAL_UEL = data['URL_QUAL_UEL']
url_UEL = data['URL_UEL']
url_QUAL_UECL = data['URL_QUAL_UECL']
url_UECL = data['URL_UECL']
url_FCWC = data['URL_FCWC']

Tabelas = []
temp = []
temp1 = []
Jogos = []
Jogos_final = []
# var_dump.var_dump(dt)
tabelas = 0

if tabelas == 1:
	# Start program 0
	result0 = subprocess.run(['python', 'TABELA_CAMPEONATOS.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	if result0.returncode != 0:
		print(result0.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result0.stdout != '':
			print('Tabela dos campeonatos criada/atualizada')
			# variavel = result0.stdout.split('\n')
			# for o in variavel:
				# Tabelas.append(o)
		else:
			print('Erro provável no subprograma TABELA_CAMPEONATOS.py')
			# var_dump.var_dump(result.stdout)



# connect to the database
conn = sqlite3.connect('championships.db')

# create a cursor object
cursor = conn.cursor()

if data0['FACS'] == 1:
	# Start first program
	result = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_FACS], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)


	# Wait for the program to finish
	if result.returncode != 0:
		print(result.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result.stdout != '':
			print('Data inicial do bolao capturada')
			# var_dump.var_dump(result.stdout)
			variavel = result.stdout.split('\n')
			while "hora ainda nao definida, setada como padrão, meio dia" in variavel:
				variavel.remove("hora ainda nao definida, setada como padrão, meio dia")
				print("hora ainda nao definida, setada como padrão, meio dia")
			while "" in variavel:
				variavel.remove("")
			inicio_bolao = variavel[0]
			print(inicio_bolao)
			inicio_bolao_date_ = datetime.strptime(inicio_bolao, "%Y-%m-%d %H:%M:%S%z")
			inicio_bolao_date = inicio_bolao_date_ - timedelta(days=4)


			if len(variavel) > 1:
				print(variavel[1])
				print('Jogo adicionado para FACS')
				Jogos.append(variavel[1])
			else:
				print("Nenhum jogo adicionado para FACS")


		else:
			print('Data inicial do bolao nao capturada')


if data0['PL'] == 1:
	# # Start second program

	# # pdb.set_trace()
	result2 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_PL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)


	# Wait for the program to finish
	if result2.returncode != 0:
		print(result2.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result2.stdout != '':
			temp = []
			print('Jogo adicionado para PL')
			variavel = result2.stdout.split('\n')
			while "" in variavel:
				variavel.remove("")
			for o in variavel:
				temp.append(o)

			result2_1 = subprocess.run(['python', 'filtro_PL.py', *temp], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

			variavel = result2_1.stdout.split('\n')
			while "" in variavel:
				variavel.remove("")

			sorted_Jogos = sorted(variavel, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))

			for o in sorted_Jogos:
				Jogos.append(o)

		else:
			print('Nenhum resultado para PL')
			# var_dump.var_dump(result.stdout)
	while "" in Jogos:
		Jogos.remove("")
	# print(Jogos)

if data0['SPFL'] == 1:
	# cursor.execute("SELECT [Número de jogos] FROM scottish_premiership_table_championship")
	# resultado1 = cursor.fetchone()[0]
	temp = []
	# cursor.execute("SELECT [Número de jogos] FROM scottish_premiership_table")
	# resultado = cursor.fetchone()[0]
	# if resultado < 33:
	if data1['RODADA_SPFL'] <= 33:
		# Start third program
		result3 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_SPFL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

		# Wait for the program to finish
		if result3.returncode != 0:
			print(result3.stderr)
		else:
			# Print the stdout output if the subprocess ran successfully
			if result3.stdout != '':
				print('Jogos SPFL encaminhados para o filtro')

				variavel = result3.stdout.split('\n')
				for o in variavel:
					temp.append(o)
				while "" in temp:
					temp.remove("")
				# print(temp)
				result3_1 = subprocess.run(['python', 'filtro_SPFL.py', *temp], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

				if result3.returncode != 0:
					print(result3.stderr)
				else:
					# Print the stdout output if the subprocess ran successfully
					if result3.stdout != '':

						variavel = result3_1.stdout.split('\n')
						while "" in variavel:
							variavel.remove("")
						sorted_Jogos = sorted(variavel, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
						for o in sorted_Jogos:
							Jogos.append(o)

			else:
				print('Nenhum resultado para SPFL')

	elif data1['RODADA_SPFL_CHAMPIONSHIP'] <= 38:
		temp=[]
		# Start third program
		result3 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_SPFL_championship], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

		# Wait for the program to finish
		if result3.returncode != 0:
			print(result3.stderr)
		else:
			# Print the stdout output if the subprocess ran successfully
			if result3.stdout != '':
				print('Jogos SPFL_championship encaminhados para o filtro')
				variavel = result3.stdout.split('\n')
				for o in variavel:
					temp.append(o)
				while "" in temp:
					temp.remove("")


			else:
				print('Nenhum resultado para SPFL_championship')

		# Start third program (part2)
		result3_2 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_SPFL_relegation], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

		# Wait for the program to finish
		if result3_2.returncode != 0:
			print(result3_2.stderr)
		else:
			# Print the stdout output if the subprocess ran successfully
			if result3_2.stdout != '':
				print('Jogos SPFL_relegation encaminhados para o filtro')
				variavel = result3_2.stdout.split('\n')
				for o in variavel:
					temp.append(o)
				while "" in temp:
					temp.remove("")
				# print(temp)
				# print("\n")


			else:
				print('Nenhum resultado para SPFL_relegation')
				# var_dump.var_dump(result.stdout)


		# print((temp))
		if len(temp) > 0:
			result3_1 = subprocess.run(['python', 'filtro_SPFL_championship_relegation.py', *temp], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
			variavel = result3_1.stdout.split('\n')
			while "" in variavel:
				variavel.remove("")
			sorted_Jogos = sorted(variavel, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
			for o in sorted_Jogos:
				Jogos.append(o)

		temp=[]
		variavel=[]

		# print(Jogos)

	# # Remove the else for now for testing propose
	else:
		result3 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_SPFL_PLAYOFFS], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

		# Wait for the program to finish
		if result3.returncode != 0:
			print(result3.stderr)
		else:
			# Print the stdout output if the subprocess ran successfully
			if result3.stdout != '':
				print('Jogos SPFL_playoffs adicionados')
				variavel = result3.stdout.split('\n')
				variavel = subst_P("P1", "P2", variavel)
				while "" in variavel:
					variavel.remove("")
				sorted_Jogos = sorted(variavel, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)
			else:
				print('Nenhum resultado para SPFL_playoffs')


if data0['LCH'] == 1:
	# cursor.execute("SELECT [Número de jogos] FROM eng_championship_table")
	# resultado = cursor.fetchone()[0]
	# if resultado < 46:
	if data1['RODADA_LCH'] <= 46:
		# # Start fourth program
		result4 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_LCH], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
		# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

		temp=[]
		variavel=[]
		# Wait for the program to finish
		if result4.returncode != 0:
			print(result4.stderr)
		else:
			# Print the stdout output if the subprocess ran successfully
			if result4.stdout != '':
				print('Jogo LCH encaminhado ao filtro')
				variavel = result4.stdout.split('\n')
				for o in variavel:
						temp.append(o)
				while "" in temp:
					temp.remove("")
				print(temp)
				result4_1 = subprocess.run(['python', 'filtro_LCH.py', *temp], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
				variavel = result4_1.stdout.split('\n')
				while "" in variavel:
					variavel.remove("")
				# print(variavel)
				sorted_Jogos = sorted(variavel, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)
			else:
				print('Nenhum resultado para LCH')


	else:

		# # Start fourth program
		result4 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_LCH_PLAYOFFS], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
		# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

		temp=[]
		variavel=[]
		# Wait for the program to finish
		if result4.returncode != 0:
			print(result4.stderr)
		else:
			# Print the stdout output if the subprocess ran successfully
			if result4.stdout != '':
				print('Jogo LCH_PLAYOFF adicionado')
				variavel = result4.stdout.split('\n')
				for o in variavel:
					temp.append(o)
				while "" in temp:
					temp.remove("")
				# print(temp)
				if "j1=3" in temp:
					temp.remove("j1=3")
					temp[(len(temp)-1)] = temp[(len(temp)-1)].replace("P1", "P3")
				# print(temp)
				index1=0
				temp = subst_P("P1", "P2", temp)

				# print(temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			else:
				print('Nenhum resultado para LCH_PLAYOFF')


if data0['FAC'] == 1:
	# # Start fifth program
	result5 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_FAC], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	# Wait for the program to finish
	if result5.returncode != 0:
		print(result5.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result5.stdout != '':
			zz=0
			variavel = result5.stdout.split('\n')
			for o in variavel:
				temp.append(o)
			while "" in temp:
				temp.remove("")
			# print(temp)
			if "j1=0" in temp:
				print('Jogo FAC anterior ao Round 3, nao adicionado')
				temp = []
			elif "j1=3" in temp:
				zz=2
				temp.remove("j1=3")
				print('Jogo FAC encaminhado ao filtro')
			elif "j1=1" in temp:
				temp.remove("j1=1")
				index1 = 0
				print('Jogo FAC semi-final adicionado')
				temp = subst_P("P1", "P2", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)


			elif "j1=2" in temp:
				temp.remove("j1=2")
				print('Jogo FAC final adicionado')
				temp = subst_P("P1", "P3", temp)
				temp = subst_P("P2", "P3", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)
			# print(temp)
			if len(temp) > 0 and zz == 2:
				result5_1 = subprocess.run(['python', 'filtro_FAC.py', *temp], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
					# # var_dump.var_dump(variavel[0])
					# print(result3_1.stdout)
				variavel = result5_1.stdout.split('\n')
				while "" in variavel:
					variavel.remove("")
				# print(variavel)
				sorted_Jogos = sorted(variavel, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

		else:
			print('Nenhum resultado para FAC')


if data0['EFL'] == 1:
	result6 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_EFL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	# Wait for the program to finish
	if result6.returncode != 0:
		print(result6.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result6.stdout != '':
			zz=0
			variavel = result6.stdout.split('\n')
			for o in variavel:
					temp.append(o)
			while "" in temp:
				temp.remove("")
			# print(temp)
			if "j1=0" in temp:
				print('Jogo EFL anterior as oitavas, nao adicionado')
				temp = []
			elif "j1=3" in temp:
				zz=2
				temp.remove("j1=3")
				print('Jogo EFL encaminhado ao filtro')
			elif "j1=1" in temp:
				temp.remove("j1=1")
				print('Jogo EFL semi-final adicionado')
				temp = subst_P("P1", "P2", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)


			elif "j1=2" in temp:
				temp.remove("j1=2")
				print('Jogo EFL final adicionado')
				temp = subst_P("P1", "P3", temp)
				temp = subst_P("P2", "P3", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)
			# print(temp)
			if len(temp) > 0 and zz == 2:
				result6_1 = subprocess.run(['python', 'filtro_FAC.py', *temp], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

				variavel = result6_1.stdout.split('\n')
				while "" in variavel:
					variavel.remove("")
				# print(variavel)
				sorted_Jogos = sorted(variavel, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

		else:
			print('Nenhum resultado para EFL')


if data0['SC'] == 1:
	result7 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_SC], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	# Wait for the program to finish
	if result7.returncode != 0:
		print(result7.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result7.stdout != '':
			zz=0
			variavel = result7.stdout.split('\n')
			for o in variavel:
				temp.append(o)
			while "" in temp:
				temp.remove("")
			# print(temp)
			if "j1=0" in temp:
				print('Jogo SC anterior as quartas, nao adicionado')
				temp = []
			elif "j1=3" in temp:
				zz=2
				temp.remove("j1=3")
				print('Jogos SC quartas-de-final adicionados')
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=1" in temp:
				temp.remove("j1=1")
				print('Jogo SC semi-final adicionado')

				temp = subst_P("P1", "P2", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)


			elif "j1=2" in temp:
				temp.remove("j1=2")
				print('Jogo SC final adicionado')

				temp = subst_P("P1", "P3", temp)

				temp = subst_P("P2", "P3", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

		else:
			print('Nenhum resultado para SC')


if data0['SLC'] == 1:
	result8 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_SLC], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	# Wait for the program to finish
	if result8.returncode != 0:
		print(result8.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result8.stdout != '':
			zz=0
			variavel = result8.stdout.split('\n')
			for o in variavel:
				temp.append(o)
			while "" in temp:
				temp.remove("")
			# print(temp)
			if "j1=0" in temp:
				print('Jogo SLC anterior as quartas, nao adicionado')
				temp = []
			elif "j1=3" in temp:
				zz=2
				temp.remove("j1=3")
				print('Jogos SLC quartas-de-final adicionados')
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=1" in temp:
				temp.remove("j1=1")
				index1 = 0
				print('Jogo SLC semi-final adicionado')
				temp = subst_P("P1", "P2", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)


			elif "j1=2" in temp:
				temp.remove("j1=2")
				index1=0
				print('Jogo SLC final adicionado')

				temp = subst_P("P1", "P3", temp)
				temp = subst_P("P2", "P3", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

		else:
			print('Nenhum resultado para SLC')

if data0['USC'] == 1:
	result9 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_USC], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	index1 = 0
	# Wait for the program to finish
	if result9.returncode != 0:
		print(result9.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result9.stdout != '':
			zz=0
			variavel = result9.stdout.split('\n')
			for o in variavel:
				temp.append(o)
			while "" in temp:
				temp.remove("")

			temp = subst_P("P1", "P2", temp)

			print("Jogo adicionado para USC")

			temp = after_FACS_filter(temp, inicio_bolao_date)

			sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
			for o in sorted_Jogos:
				Jogos.append(o)

		else:
			print('Nenhum resultado para USC')

if data0['QUAL_UCL'] == 1:
	result10 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_QUAL_UCL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	index1 = 0
	# Wait for the program to finish
	if result10.returncode != 0:
		print(result10.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result10.stdout != '':
			zz=0
			variavel = result10.stdout.split('\n')
			for o in variavel:
				temp.append(o)
			while "" in temp:
				temp.remove("")

			print("Jogo encaminhado ao filtro para qualificatorias UCL")

			temp = after_FACS_filter(temp, inicio_bolao_date)

			temp = subst_P("P2", "P1", temp)

			sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
			for o in sorted_Jogos:
				Jogos.append(o)

		else:
			print('Nenhum resultado para QUAL-UCL')


if data0['UCL'] == 1:
	result11 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_UCL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	# Wait for the program to finish
	if result11.returncode != 0:
		print(result11.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result11.stdout != '':
			zz=0
			variavel = result11.stdout.split('\n')
			for o in variavel:
					temp.append(o)
			while "" in temp:
				temp.remove("")
			# print(temp)
			if "j1=0" in temp:
				temp.remove("j1=0")
				print('Jogo UCL adicionado')

				# pdb.set_trace()
				temp = subst_P("P2", "P1", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=3" in temp:
				zz=2
				temp.remove("j1=3")

				temp = subst_P("P1", "P2", temp)

				print('Jogo UCL oitavas/quartas adicionado')

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=1" in temp:
				temp.remove("j1=1")
				index1 = 0
				print('Jogo UCL semi-final adicionado')
				temp = subst_P("P1", "P3", temp)
				temp = subst_P("P2", "P3", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)


			elif "j1=2" in temp:
				temp.remove("j1=2")
				index1=0
				print('Jogo UCL final adicionado')

				temp = subst_P("P1", "P4", temp)
				temp = subst_P("P2", "P4", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

		else:
			print('Nenhum resultado para UCL')



if data0['QUAL_UEL'] == 1:
	result12 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_QUAL_UEL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	index1 = 0
	# Wait for the program to finish
	if result12.returncode != 0:
		print(result12.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result12.stdout != '':
			zz=0
			variavel = result12.stdout.split('\n')
			for o in variavel:
				temp.append(o)
			while "" in temp:
				temp.remove("")

			print("Jogo encaminhado ao filtro para qualificatorias UEL")

			temp = after_FACS_filter(temp, inicio_bolao_date)

			if len(temp) > 0:
				temp = subst_P("P2", "P1", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)
		else:
			print('Nenhum resultado para QUAL-UEL')



if data0['UEL'] == 1:
	result13 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_UEL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]
	# Wait for the program to finish
	# pdb.set_trace()
	if result13.returncode != 0:
		print(result13.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result13.stdout != '':
			zz=0
			variavel = result13.stdout.split('\n')
			for o in variavel:
					temp.append(o)
			while "" in temp:
				temp.remove("")
			# print(temp)
			if "j1=0" in temp:
				temp.remove("j1=0")
				print('Jogo UEL adicionado')

				# pdb.set_trace()
				temp = subst_P("P2", "P1", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=3" in temp:
				zz=2
				temp.remove("j1=3")

				temp = subst_P("P1", "P2", temp)

				print('Jogo UEL oitavas/quartas adicionado')

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=1" in temp:
				temp.remove("j1=1")
				index1 = 0
				print('Jogo UEL semi-final adicionado')
				temp = subst_P("P1", "P3", temp)
				temp = subst_P("P2", "P3", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)


			elif "j1=2" in temp:
				temp.remove("j1=2")
				index1=0
				print('Jogo UEL final adicionado')

				temp = subst_P("P1", "P4", temp)
				temp = subst_P("P2", "P4", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

		else:
			print('Nenhum resultado para UEL')

if data0['QUAL_UECL'] == 1:
	result14 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_QUAL_UECL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

	temp=[]
	variavel=[]
	index1 = 0
	# Wait for the program to finish
	if result14.returncode != 0:
		print(result14.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result14.stdout != '':
			zz=0
			variavel = result14.stdout.split('\n')
			for o in variavel:
				temp.append(o)
			while "" in temp:
				temp.remove("")

			print("Jogo encaminhado ao filtro para qualificatorias UECL")

			temp = after_FACS_filter(temp, inicio_bolao_date)

			if len(temp) > 0:
				temp = subst_P("P2", "P1", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)
		else:
			print('Nenhum resultado para QUAL-UECL')



if data0['UECL'] == 1:
	result15 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_UECL], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
	# # p1 = subprocess.Popen(['python', 'iniciador_de_string.py'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	temp=[]
	variavel=[]

	# Wait for the program to finish
	if result15.returncode != 0:
		print(result15.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result15.stdout != '':
			zz=0
			variavel = result15.stdout.split('\n')
			for o in variavel:
					temp.append(o)
			while "" in temp:
				temp.remove("")
			# print(temp)
			if "j1=0" in temp:
				temp.remove("j1=0")
				print('Jogo UECL adicionado')

				# pdb.set_trace()
				temp = subst_P("P2", "P1", temp)


				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=3" in temp:
				zz=2
				temp.remove("j1=3")

				temp = subst_P("P2", "P1", temp)

				print('Jogo UECL oitavas/quartas adicionado')

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=1" in temp:
				temp.remove("j1=1")
				index1 = 0
				print('Jogo UECL semi-final adicionado')
				temp = subst_P("P1", "P2", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)



			elif "j1=2" in temp:
				temp.remove("j1=2")
				index1=0
				print('Jogo UECL final adicionado')

				temp = subst_P("P1", "P3", temp)
				temp = subst_P("P2", "P3", temp)

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

		else:
			print('Nenhum resultado para UECL')



if data0['FCWC'] == 1:
	result16 = subprocess.run(['python', 'buscador_resultados.py', *dt, *dt2, url_FCWC], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

	temp=[]
	variavel=[]
	# Wait for the program to finish
	if result16.returncode != 0:
		print(result16.stderr)
	else:
		# Print the stdout output if the subprocess ran successfully
		if result16.stdout != '':
			zz=0
			variavel = result16.stdout.split('\n')
			for o in variavel:
				temp.append(o)
			while "" in temp:
				temp.remove("")
			# print(temp)
			if "j1=0" in temp:
				temp.remove("j1=0")
				print('Jogo anterior a semi FCWC')

				temp = subst_P("P2", "P1", temp)


				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=3" in temp:
				zz=2
				temp.remove("j1=3")

				temp = subst_P("P2", "P1", temp)

				print('Jogo FCWC semi adicionado')

				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

			elif "j1=1" in temp:
				temp.remove("j1=1")
				index1 = 0
				print('Jogo FCWC terceiro adicionado')
				temp = subst_P("P2", "P1", temp)
				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)


			elif "j1=2" in temp:
				temp.remove("j1=2")
				index1=0
				print('Jogo FCWC final adicionado')

				temp = subst_P("P1", "P2", temp)


				sorted_Jogos = sorted(temp, key=lambda s: tuple(s.split(" zzz")[0].split(' - ')[-2:]))
				for o in sorted_Jogos:
					Jogos.append(o)

		else:
			print('Nenhum resultado para FCWC')


# print(Jogos)
for o in range(len(Jogos)):
	Jogos[o] = Jogos[o].replace("_QUAL", '')
	print(Jogos[o])


# print(Jogos)

if data0['RESULTADOS_NO_EXCEL'] == 1:

	workbook_name = data['ARQUIVO_EXCEL_ABRIR_2']
	workbook_future_name = data['ARQUIVO_EXCEL_SALVAR']

	placares = []
	for index, o in enumerate(Jogos):
		# z=Jogos[index].split("zzz")
		# print(z[1])
		if "aet" in Jogos[index].split("zzz")[1]:
			placares.append(Jogos[index].split("zzz")[1].split(" (")[0].split(":")[0])
			placares.append(Jogos[index].split("zzz")[1].split(" (")[0].split(":")[1])
		if "pso" in Jogos[index].split("zzz")[1]:
			placares.append(Jogos[index].split("zzz")[1].split(") ")[0].split(", ", 2)[1].split(":")[0])
			placares.append(Jogos[index].split("zzz")[1].split(") ")[0].split(", ", 2)[1].split(":")[1])
		else:
			placares.append(Jogos[index].split("zzz")[1].split(" (")[0].split(":")[0])
			placares.append(Jogos[index].split("zzz")[1].split(" (")[0].split(":")[1])

	lista = []
	for i in range(len(placares)):
		temp = []
		if i % 2 == 0:
			temp.extend([int(placares[i]), "x", int(placares[i+1])])
			lista.append(temp)

		else:
			continue

	workbook = load_workbook(filename = workbook_name, read_only=False,keep_vba=True)
	sheet = workbook['Tabela']
	for x in range(5000):
		if sheet[('A'+str(x+1))].value == data['Nome_Rodada']:
			position = ('A'+str(x+1))

	start_row = (int(position[1:])+1)
	start_column = openpyxl.utils.column_index_from_string("B")

	for i, value in enumerate(lista):
		for i1 in range(len(lista[i])):

			cell = sheet.cell(row=(start_row+i), column=(start_column+i1))
			cell.value = lista[i][i1]

	workbook.save(workbook_future_name)